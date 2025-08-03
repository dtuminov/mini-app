# Aiogram imports
# Another
import os
import shutil

from aiogram import Dispatcher, types, filters, Bot, F
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import ContentType
from openpyxl import load_workbook

# DB
from database.orm import ORM as orm
# Keyboards
from keyboards.admin_su_kbs import *
# States
from states.main_states import AdminStates


# loader
def load_admin_search_user_handler(dispatcher: Dispatcher, bot: Bot):

	# init
	global g_bot, g_storage
	g_bot = bot
	g_storage = dispatcher.storage

	dispatcher.message.register(
		search_user_enterEntity,
		F.text == "🔎 Поиск пользователя",
		StateFilter(AdminStates.main_menu)
	)

	dispatcher.message.register(
		search_user_getUser_msg,
		F.content_type == ContentType.TEXT,
		StateFilter(AdminStates.search_user_enterEntity)
	)

	# Back handler
	dispatcher.callback_query.register(
		search_user_getUser_query,
		F.data.startswith("back_to_su_"),
		StateFilter(
			AdminStates.search_user_actions,
			AdminStates.search_user_enterAttempts,
			AdminStates.search_user_enterFreeze,
			AdminStates.search_user_enterCard
		)
	)

	# Attempts handlers
	dispatcher.callback_query.register(
		search_user_addAttempts_1,
		F.data.startswith("su_edit_attempts_"),
		StateFilter(AdminStates.search_user_actions)
	)

	dispatcher.message.register(
		search_user_addAttempts_2,
		F.content_type == ContentType.TEXT,
		StateFilter(AdminStates.search_user_enterAttempts)
	)

	# Freeze handlers
	dispatcher.callback_query.register(
		search_user_addFreeze_1,
		F.data.startswith("su_edit_freeze_"),
		StateFilter(AdminStates.search_user_actions)
	)

	dispatcher.message.register(
		search_user_addFreeze_2,
		F.content_type == ContentType.TEXT,
		StateFilter(AdminStates.search_user_enterFreeze)
	)

	# Card handlers
	dispatcher.callback_query.register(
		search_user_addCard_1,
		F.data.startswith("su_add_card_"),
		StateFilter(AdminStates.search_user_actions)
	)

	dispatcher.message.register(
		search_user_addCard_2,
		F.content_type == ContentType.TEXT,
		StateFilter(AdminStates.search_user_enterCard)
	)

	# Delete card handlers
	dispatcher.callback_query.register(
		search_user_delCard_1,
		F.data.startswith("su_del_card_"),
		StateFilter(AdminStates.search_user_actions)
	)

	dispatcher.message.register(
		search_user_delCard_2,
		F.content_type == ContentType.TEXT,
		StateFilter(AdminStates.search_user_enterdelCard)
	)

	# Transactions list handler
	dispatcher.callback_query.register(
		process_transactions_list,
		F.data.startswith("su_transactions_"),
		StateFilter("*")
	)

	
async def search_user_enterEntity(message: types.Message, state: FSMContext):
	
	msg_text = "Введите ID пользователя или его Username:"

	await message.answer(
		text = msg_text,
		reply_markup = get_back_kb()
	)

	await AdminStates.search_user_enterEntity.set()


async def search_user_getUser_msg(message: types.Message, state: FSMContext):

	if message.text.startswith("@"):
		user = await orm.get_user_byUsername(message.text)

		if user is None:
			await message.answer(
				text = "🔴 Пользователь не найден! Повторите попытку:"
			)
			return
		
	else:
		if not message.text.replace("id", "").isdigit():
			await message.answer("🔴 Формат: idЦИФРЫ или ЦИФРЫ или @USERNAME! Повторите попытку:")
			return
		
		user = await orm.get_user(int(message.text.replace("id", "")))

		if user is None:
			await message.answer(
				text = "🔴 Пользователь не найден! Повторите попытку:"
			)
			return
		
	msg_text = f"""👤 Пользователь id{user.user_id} ({user.username})

👨‍💻 Полное имя: {user.fullname}
📆 Дата первого запуска: {user.register_date.strftime(r"%d-%m-%y %H:%M:%S")} (UTC)
🌐 Всего рейтинга: {user.points}
🍁 Сезонного рейтинга: {user.season_points}
🔎 Попыток: {user.attempts}
☃️ Заморозка: {user.freeze}ч"""

	await message.answer(msg_text, reply_markup = get_actions_ikb(user.user_id))

	await AdminStates.search_user_actions.set()


async def search_user_getUser_query(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])
	user = await orm.get_user(user_id)

	msg_text = f"""👤 Пользователь id{user.user_id} ({user.username})

👨‍💻 Полное имя: {user.fullname}
📆 Дата первого запуска: {user.register_date.strftime(r"%d-%m-%y %H:%M:%S")} (UTC)
🌐 Всего рейтинга: {user.points}
🍁 Сезонного рейтинга: {user.season_points}
🔎 Попыток: {user.attempts}
☃️ Заморозка: {user.freeze}ч"""

	await query.message.edit_text(msg_text, reply_markup = get_actions_ikb(user.user_id))

	await AdminStates.search_user_actions.set()

	await g_bot.answer_callback_query(query.id)


# add attempts
async def search_user_addAttempts_1(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])

	await query.message.edit_text(
		text = "Введите число попыток (напр. 1 или -1):",
		reply_markup = get_back_ikb(user_id = user_id)
	)

	async with state.proxy() as storage:
		storage["su_aa_uid"] = user_id
		storage["su_aa_lmi"] = query.message.message_id

	await AdminStates.search_user_enterAttempts.set()

	await g_bot.answer_callback_query(query.id)


async def search_user_addAttempts_2(message: types.Message, state: FSMContext):

	# unpack
	async with state.proxy() as storage:
		uid = storage.get("su_aa_uid")
		lmi = storage.get("su_aa_lmi")

	try:
		await g_bot.edit_message_reply_markup(
			message.chat.id,
			lmi,
			None
		)
	except:
		pass

	# validation
	try:
		kf = int(message.text)
	except:
		await message.answer(
			text = "🔴 Ошибка! Повторите попытку:",
			reply_markup = get_back_ikb(uid)
		)
		return
	
	user = await orm.get_user(uid)
	await orm.set_users_field(user.user_id, "attempts", user.attempts + kf)

	# notify
	try:
		await g_bot.send_message(
			chat_id = user.user_id,
			text = f"🎁 Вам было начислено {kf} попыток!" if kf > 0 else f"💔 С вас сняли {kf} попыток..."
		)
	except:
		pass

	await message.answer(
		text = "✅ Готово!",
		reply_markup = get_back_ikb(uid)
	)

	await AdminStates.search_user_actions.set()


# add attempts
async def search_user_addFreeze_1(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])

	await query.message.edit_text(
		text = "Введите число часов заморозки (напр. 1 или -1):",
		reply_markup = get_back_ikb(user_id = user_id)
	)

	async with state.proxy() as storage:
		storage["su_af_uid"] = user_id
		storage["su_af_lmi"] = query.message.message_id

	await AdminStates.search_user_enterFreeze.set()

	await g_bot.answer_callback_query(query.id)


async def search_user_addFreeze_2(message: types.Message, state: FSMContext):

	# unpack
	async with state.proxy() as storage:
		uid = storage.get("su_af_uid")
		lmi = storage.get("su_af_lmi")

	try:
		await g_bot.edit_message_reply_markup(
			message.chat.id,
			lmi,
			None
		)
	except:
		pass

	# validation
	try:
		kf = int(message.text)
	except:
		await message.answer(
			text = "🔴 Ошибка! Повторите попытку:",
			reply_markup = get_back_ikb(uid)
		)
		return
	
	user = await orm.get_user(uid)
	await orm.set_users_field(user.user_id, "freeze", user.freeze + kf)

	# notify
	try:
		await g_bot.send_message(
			chat_id = user.user_id,
			text = f"☃️ Вы были заморожены на {kf}ч!" if kf > 0 else f"☃️ Вы были разморожены на {kf}ч!"
		)
	except:
		pass

	await message.answer(
		text = "✅ Готово!",
		reply_markup = get_back_ikb(uid)
	)

	await AdminStates.search_user_actions.set()


# add card
async def search_user_addCard_1(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])

	await query.message.edit_text(
		text = "Введите ID карты (напр. #1 или 1):",
		reply_markup = get_back_ikb(user_id = user_id)
	)

	async with state.proxy() as storage:
		storage["su_ac_uid"] = user_id
		storage["su_ac_lmi"] = query.message.message_id

	await AdminStates.search_user_enterCard.set()

	await g_bot.answer_callback_query(query.id)


async def search_user_addCard_2(message: types.Message, state: FSMContext):

	# unpack
	async with state.proxy() as storage:
		uid = storage.get("su_ac_uid")
		lmi = storage.get("su_ac_lmi")

	try:
		await g_bot.edit_message_reply_markup(
			message.chat.id,
			lmi,
			None
		)
	except:
		pass

	# validation
	try:
		card_id = int(message.text.replace("#", ""))

		card = await orm.get_card_by_cardID(card_id)

		if card is None:
			raise
	except:
		await message.answer(
			text = "🔴 Ошибка! Повторите попытку:",
			reply_markup = get_back_ikb(uid)
		)
		return
	
	user = await orm.get_user(uid)
	user.inventory.append(card_id)
	await orm.set_users_field(user.user_id, "inventory", user.inventory)

	user.points += card.card_weight
	user.season_points += card.card_weight

	await orm.set_users_field(user.user_id, "points",  user.points)
	await orm.set_users_field(user.user_id, "season_points", user.season_points)
	await orm.set_users_field(user.user_id, "stream_points", user.stream_points + card.card_weight)

	# notify
	try:
		with open(card.card_image, "rb") as image_bin:
			await g_bot.send_photo(
				photo = image_bin,
				chat_id = user.user_id,
				caption = f"🎁 Вам подарили <b>{card.card_name} ({card.card_rarity.title()})</b>!\n🌐 +{card.card_weight} Онлайна\n\nОнлайн за все время: {user.points}\nОнлайн за текущий стрим: {user.season_points}"
			)
	except:
		pass

	await message.answer(
		text = "✅ Готово!",
		reply_markup = get_back_ikb(uid)
	)

	await AdminStates.search_user_actions.set()


# add card
async def search_user_delCard_1(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])

	await query.message.edit_text(
		text = "Введите ID карты для удаления (напр. #1 или 1):",
		reply_markup = get_back_ikb(user_id = user_id)
	)

	async with state.proxy() as storage:
		storage["su_ac_uid"] = user_id
		storage["su_ac_lmi"] = query.message.message_id

	await AdminStates.search_user_enterdelCard.set()

	await g_bot.answer_callback_query(query.id)


async def search_user_delCard_2(message: types.Message, state: FSMContext):

	# unpack
	async with state.proxy() as storage:
		uid = storage.get("su_ac_uid")
		lmi = storage.get("su_ac_lmi")

	try:
		await g_bot.edit_message_reply_markup(
			message.chat.id,
			lmi,
			None
		)
	except:
		pass

	# validation
	try:
		card_id = int(message.text.replace("#", ""))

		card = await orm.get_card_by_cardID(card_id)

		if card is None:
			raise
	except:
		await message.answer(
			text = "🔴 Ошибка! Повторите попытку:",
			reply_markup = get_back_ikb(uid)
		)
		return
	
	user = await orm.get_user(uid)

	if card_id not in user.inventory:
		await message.answer(
			text = "⚠️ Указанной карты нет у пользователя!",
			reply_markup = get_back_ikb(uid)
		)
		return

	user.inventory.remove(card_id)
	await orm.set_users_field(user.user_id, "inventory", user.inventory)

	user.points -= card.card_weight
	user.season_points -= card.card_weight

	await orm.set_users_field(user.user_id, "points",  user.points)
	await orm.set_users_field(user.user_id, "season_points", user.season_points)
	await orm.set_users_field(user.user_id, "stream_points", user.stream_points + card.card_weight)

	await message.answer(
		text = "✅ Готово!",
		reply_markup = get_back_ikb(uid)
	)

	await AdminStates.search_user_actions.set()


async def process_transactions_list(query: types.CallbackQuery, state: FSMContext):

	user_id = int(query.data.split("_")[-1])

	# copy the table
	table_path = shutil.copy(src = "templates/transactions.xlsx", dst = f"templates/transactions_{user_id}.xlsx")

	# load table
	book = load_workbook(filename = table_path)
	sheet = book["users"]

	# get all users
	all_transactions = (await orm.get_user_transactions(user_id))[:1000]
	all_transactions.sort(reverse = True, key = lambda x: x.transaction_id)

	# write
	for row, trans in enumerate(all_transactions, 2):
		# write row
		sheet.cell(row = row, column = 1, value = trans.transaction_id)
		sheet.cell(row = row, column = 2, value = trans.transaction_type)
		sheet.cell(row = row, column = 3, value = trans.transaction_date)
		sheet.cell(row = row, column = 4, value = ", ".join(trans.transaction_in))
		sheet.cell(row = row, column = 5, value = ", ".join(trans.transaction_out))

	# save book
	book.save(table_path)

	# answer
	with open(file = table_path, mode = "rb") as table:
		await query.message.answer_document(document = table)

	# delete temp table
	try:
		os.remove(table_path)
	except:
		pass
	
	await g_bot.answer_callback_query(query.id)